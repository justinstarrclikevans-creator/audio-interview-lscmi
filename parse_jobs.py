#!/usr/bin/env python3
"""
Parse jobs from all available .xlsx / .csv files in Facilitation Scoring and data folders.
"""

import sys
import json
import os
import openpyxl

def parse_jobs_spreadsheets():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    search_dirs = [
        os.path.join(base_dir, '..', 'Facilitation Scoring'),
        os.path.join(base_dir, 'data'),
        os.path.join(base_dir, '..')
    ]

    found_files = []
    for d in search_dirs:
        if os.path.exists(d):
            for f in os.listdir(d):
                if ('job' in f.lower() or 'employer' in f.lower()) and f.endswith('.xlsx'):
                    found_files.append(os.path.join(d, f))

    jobs = []
    seen = set()

    for file_path in found_files:
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                if 'research' in sheet_name.lower() or 'note' in sheet_name.lower():
                    continue
                sheet = wb[sheet_name]
                for idx, row in enumerate(sheet.iter_rows(values_only=True)):
                    if idx == 0 or not any(row):
                        continue
                    
                    company = str(row[0]).strip() if len(row) > 0 and row[0] else 'Local Employer'
                    job_title = str(row[1]).strip() if len(row) > 1 and row[1] else 'Specialist'
                    location = str(row[2]).strip() if len(row) > 2 and row[2] else 'Charleston, SC'
                    pay_rate = str(row[3]).strip() if len(row) > 3 and row[3] else 'Competitive'
                    description = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                    careers_url = str(row[6]).strip() if len(row) > 6 and row[6] else ''

                    if pay_rate.lower() in ['none', 'not disclosed', 'null']:
                        pay_rate = 'Competitive / Market Rate'

                    key = (company.lower(), job_title.lower(), location.lower())
                    if key in seen:
                        continue
                    seen.add(key)

                    jobs.append({
                        'company': company,
                        'jobTitle': job_title,
                        'location': location,
                        'payRate': pay_rate,
                        'description': description,
                        'careersUrl': careers_url,
                        'sourceFile': os.path.basename(file_path)
                    })
        except Exception as e:
            sys.stderr.write(f"Error reading {file_path}: {e}\n")

    return jobs

if __name__ == '__main__':
    all_jobs = parse_jobs_spreadsheets()
    print(json.dumps(all_jobs))
